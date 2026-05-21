"""Generación XLSX de conciliación (SCRUM-190)."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.schemas.conciliacion import (
    ConciliacionDetalleIncapacidadItem,
    ConciliacionResumenEntidadItem,
)


def _autosize_columns(ws) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 48)


def _escribir_hoja_resumen(
    wb: Workbook,
    resumenes: list[ConciliacionResumenEntidadItem],
) -> None:
    ws = wb.active
    ws.title = "Resumen"
    headers = [
        "Entidad",
        "Total cobrado",
        "Total pagado",
        "Diferencia",
        "Cobradas en periodo",
        "Pendientes de pago",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in resumenes:
        ws.append(
            [
                row.entidad,
                float(row.total_cobrado),
                float(row.total_pagado),
                float(row.diferencia),
                row.cantidad_cobrada_periodo,
                row.cantidad_pendiente_pago,
            ]
        )
    _autosize_columns(ws)


def _escribir_hoja_detalle(
    wb: Workbook,
    detalle: list[ConciliacionDetalleIncapacidadItem],
) -> None:
    ws = wb.create_sheet("Detalle")
    headers = [
        "Radicado",
        "Estado",
        "Colaborador",
        "Entidad",
        "Tipo incapacidad",
        "Fecha recepción",
        "Liquidado",
        "Monto pagado",
        "Referencia pago",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for item in detalle:
        ws.append(
            [
                item.radicado,
                item.estado,
                item.colaborador_nombre or "",
                item.entidad_nombre or "",
                item.incapacidad_tipo_extraido or "",
                item.fecha_recepcion.isoformat(),
                "Sí" if item.liquidado else "No",
                float(item.monto_pagado) if item.monto_pagado is not None else "",
                item.referencia_pago or "",
            ]
        )
    _autosize_columns(ws)


def generar_xlsx_conciliacion(
    *,
    resumenes: list[ConciliacionResumenEntidadItem],
    detalle: list[ConciliacionDetalleIncapacidadItem],
) -> bytes:
    wb = Workbook()
    _escribir_hoja_resumen(wb, resumenes)
    _escribir_hoja_detalle(wb, detalle)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

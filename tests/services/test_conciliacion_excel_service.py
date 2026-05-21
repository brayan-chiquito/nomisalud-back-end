"""Tests de generación XLSX de conciliación."""

import uuid
from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.schemas.conciliacion import (
    ConciliacionDetalleIncapacidadItem,
    ConciliacionResumenEntidadItem,
)
from app.services.conciliacion_excel_service import generar_xlsx_conciliacion


def test_generar_xlsx_hojas_resumen_y_detalle():
    resumen = [
        ConciliacionResumenEntidadItem(
            entidad="NomiSalud",
            total_cobrado=Decimal("1000.00"),
            total_pagado=Decimal("800.00"),
            diferencia=Decimal("200.00"),
            cantidad_cobrada_periodo=2,
            cantidad_pendiente_pago=1,
        )
    ]
    detalle = [
        ConciliacionDetalleIncapacidadItem(
            id=uuid.UUID("00000000-0000-0000-0000-000000000001"),
            radicado="IN01",
            estado="cobrada",
            colaborador_nombre="Ana",
            entidad_nombre="NomiSalud",
            incapacidad_tipo_extraido="enfermedad",
            fecha_recepcion=datetime(2024, 5, 10, tzinfo=UTC),
            monto_pagado=None,
            referencia_pago=None,
            liquidado=False,
        )
    ]
    data = generar_xlsx_conciliacion(resumenes=resumen, detalle=detalle)
    wb = load_workbook(filename=BytesIO(data))
    assert wb.sheetnames == ["Resumen", "Detalle"]
    assert wb["Resumen"]["A2"].value == "NomiSalud"
    assert wb["Detalle"]["A2"].value == "IN01"
